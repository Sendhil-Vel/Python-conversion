import openpyxl
import pandas as pd
import xlwt
from xlwt import Workbook
from process import logging
from process import env


processList = []

def FinalList(itemData,headerlist,filename) :
    print("***********Generating File " + env.SourceFolder+"\\" + filename)
    finalist = []
    tmpFst = []
    tmpFst = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17]
    finalist.append(tmpFst)
    for data in itemData :
        tmpFst = []
        for hData in headerlist :
            tmpFst.append(hData)
        for dData in data :
            tmpFst.append(dData)
        finalist.append(tmpFst)  
    fileWrite(finalist,filename)
    filename = str.replace(filename,".xlsx",".xls")
    filename = str.replace(filename,"\\xlsx\\","\\xls\\")
    print("***********Generating File " + env.SourceFolder+"\\" + filename)
    fileWritexls(finalist,filename)

def fileWritexls(finalDataObj,filename) :
    wb = Workbook()
    sheet1 = wb.add_sheet('Sheet 1')
    rowcnt = -1
    colcnt = -1
    for rowdata in finalDataObj:
        rowcnt = rowcnt + 1
        colcnt = -1
        for coldata in rowdata :
            colcnt = colcnt  + 1
            sheet1.write(rowcnt,colcnt,coldata)
    wb.save(filename)
        
def fileWrite(finalDataObj,filename) :
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in finalDataObj:
        sheet.append(row)
    workbook.save(filename)
    
def processItem(itemData) :
    item = []
    item.append(itemData[1])
    item.append(itemData[2])
    item.append(itemData[3])
    item.append(itemData[4])
    item.append(str(itemData[7]))
    item.append(str(itemData[8]))
    item.append(itemData[9])
    item.append(itemData[10])
    return item
    
def processHeader(rawData) :
    headList = []
    headList.append((str.replace(str.replace(str(rawData[7][1]),"SUPPLER CODE",""),":","")).strip()+" "+str(rawData[7][4]))
    headList.append(rawData[8][1])
    headList.append((str.replace(str.replace(str(rawData[10][1]),"DELIVERY DATE",""),":","")).strip())
    headList.append((str.replace(str.replace(str(rawData[11][1]),"OUR IN CHARGE",""),":","")).strip())
    data = str.split(rawData[12][1],"(")
    headList.append((str.replace(str.replace(str(data[0]),"CURRENCY",""),":","")).strip())
    headList.append("("+data[1])
    headList.append((str.replace(str.replace(str(rawData[8][7]),"P/O NO.",""),":","")).strip())
    headList.append((str.replace(str.replace(str(rawData[9][7]),"DATE",""),":","")).strip())
    headList.append((str.replace(str.replace(str(rawData[10][7]),"PROJECT#",""),":","")).strip())
    return headList

def processData(rawData) :
    headList = processHeader(rawData)
    startItem = False
    startFooter = False
    subLists = []
    for item in rawData :
        if item[1] == "ACC" :
            startItem = True
        if (str.replace(str.replace(str(item[1]),"/",""),"<","")).find("TOTAL") > -1  or (str(item[9])).find("TOTAL") > -1 :
            startFooter = True
            startItem = False
        if startItem == True :
            if ((str(item[2])).strip()).isnumeric() == True :
                rowData = processItem(item)
                subLists.append(rowData)
            else :
                if (str(item[1]) == "None" and str(item[2]) == "None" and str(item[3]) == "None" and str(item[4]) != "None") :
                    subLists[len(subLists)-1][3] = subLists[len(subLists)-1][3] + " " + str(item[4])
    return subLists,headList
        
def processXlsFiles(file,xlsFile) :
    dataframe = openpyxl.load_workbook(file)
    dataframe1 = dataframe.active
    cnt = 0
    fList = []
    HeadLists = []
    for row in range(0, dataframe1.max_row):
        subList = []
        cnt = cnt + 1
        subList.append(cnt)
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            subList.append(str(col[row].value))
        fList.append(subList)
    subLists, HeadLists = processData(fList)
    FinalList(subLists,HeadLists,xlsFile)
    